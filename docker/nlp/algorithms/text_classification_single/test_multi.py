# 用于多标签测试集预测，并按“命中任一真实标签即正确”的规则统计指标。

import os
import json
import re
from argparse import ArgumentParser
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd
import torch

from classifier_model import BertForMultiLabel
from transformers import BertTokenizer


default_label_file_path = "/user_home/du.jing/国家安全/单标签/data/labels.json"
default_test_data_path = "./data/test_0709_top11_labels.json"
SCRIPT_DIR = Path(__file__).resolve().parent

TRUE_LABEL_FIELD = "labels"
PRED_LABEL_FIELD = "model_pred"
PRED_PROB_FIELD = "model_prob"
TOP_GT_LABEL_RATIO = 0.5
OVERALL_COMPARE_XLSX = "overall_metrics_summary.xlsx"

tokenizer = None
model = None
label2int = None
int2label = None


def load_model(model_file_path, label_file_path):
    global tokenizer
    global model
    global label2int
    global int2label

    GPU_ID = 1
    device = torch.device(f"cuda:{GPU_ID}" if torch.cuda.is_available() else "cpu")
    print(f"当前使用的设备: {device}")

    tokenizer = BertTokenizer.from_pretrained(model_file_path)
    model = BertForMultiLabel.from_pretrained(model_file_path)
    model.to(device)
    model.eval()

    with open(label_file_path, "r", encoding="utf-8") as f:
        label2int = json.load(f)
    int2label = {v: k for k, v in label2int.items()}
    print(int2label)


def pre_process_text(text: str) -> str:
    text = re.sub(r"<[a-zA-Z/].*?>", "", text)
    text = re.sub(r"http[\x00-\x1f\x21-\xff]+", "", text)
    text = re.sub(r"www\.[\x00-\x1f\x21-\xff]+", "", text)
    text = re.sub(r"&nbsp;|&quot;|\u200b|\u3000", "", text)
    return text


def token2id(text, padding=True):
    inputs = tokenizer(text, max_length=512, padding="max_length")
    if len(inputs["input_ids"]) > 512:
        input_ids = [inputs["input_ids"][:256] + inputs["input_ids"][-256:]]
        attention_mask = [inputs["attention_mask"][:512]]
    else:
        if padding:
            input_ids = [inputs["input_ids"]]
            attention_mask = [inputs["attention_mask"]]
        else:
            input_ids = [[i for i in inputs["input_ids"] if i > 0]]
            attention_mask = [[i for i in inputs["attention_mask"] if i > 0]]

    return input_ids, attention_mask


def predict(input_ids, attention_mask):
    output = model(
        input_ids=torch.tensor(input_ids).to(model.device),
        attention_mask=torch.tensor(attention_mask).to(model.device),
    )

    logits = output[0]
    loss_type = getattr(model.config, "loss_type", None)
    if loss_type in ["BCE", "ASL"]:
        probs = torch.sigmoid(logits)
    else:
        probs = torch.softmax(logits, dim=-1)

    pred_id = torch.argmax(probs, dim=-1).item()
    pred_prob = probs[0][pred_id].item()
    pred_label = int2label[pred_id]

    label_list = pred_label
    prob_list = pred_prob

    return label_list, prob_list


def run_warmup():
    text = "资产配置2.0版本（202606） 1、QQQ-20%，721 2、SMH-15%，620 3、CRCL-15%，78 4、科创50-10%，1.756 5、生息类-10%（茅台1290/国债/长江电力28） 6、GOOGL-5%，360 7、SPACX-5%，160 8、期货/合约/期权-8% 9、BTC-3%，63000 9、黄金-3%，4226 10、BNB-3% ，59"
    text = pre_process_text(text)
    input_ids, attention_mask = token2id(text)
    pred_label, pred_prob = predict(input_ids, attention_mask)
    print(pred_prob)
    print(pred_label)


def read_json_or_jsonl(file_path: str) -> List[Dict[str, Any]]:
    """
    兼容读取：
    1. jsonl：每行一个 JSON 字典
    2. json：标准 JSON 列表文件
    3. json：单个 JSON 字典文件
    4. 即使后缀写错，也会自动尝试整体 JSON / JSONL 两种方式
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    data = []

    with open(file_path, "r", encoding="utf-8") as f:
        text = f.read().strip()

    if not text:
        print(f"[警告] 文件为空: {file_path}")
        return data

    try:
        raw = json.loads(text)

        if isinstance(raw, list):
            for idx, item in enumerate(raw):
                if not isinstance(item, dict):
                    print(f"[格式错误] 文件: {file_path}, JSON列表索引: {idx}, 不是字典: {type(item)}")
                    continue
                data.append(item)
            return data

        if isinstance(raw, dict):
            data.append(raw)
            return data

        raise ValueError(f"JSON根对象既不是列表也不是字典，实际类型: {type(raw)}")

    except Exception:
        pass

    with open(file_path, "r", encoding="utf-8") as f:
        for line_idx, line in enumerate(f, start=1):
            line = line.strip()
            if not line:
                continue

            try:
                item = json.loads(line)
            except Exception as e:
                print(f"[JSONL解析失败] 文件: {file_path}, 行号: {line_idx}, 错误: {repr(e)}")
                continue

            if not isinstance(item, dict):
                print(f"[格式错误] 文件: {file_path}, 行号: {line_idx}, 不是字典: {type(item)}")
                continue

            data.append(item)

    return data


def str_to_bool(value):
    if isinstance(value, bool):
        return value
    value = str(value).strip().lower()
    if value in {"1", "true", "yes", "y", "on"}:
        return True
    if value in {"0", "false", "no", "n", "off"}:
        return False
    raise ValueError(f"Invalid boolean value: {value}")


def normalize_label(label: Any) -> Optional[str]:
    if label is None:
        return None
    label = str(label).strip()
    return label or None


def get_label_list(value: Any) -> List[str]:
    if value is None:
        return []

    raw_labels = value if isinstance(value, list) else [value]
    labels = []
    seen = set()
    for item in raw_labels:
        label = normalize_label(item)
        if label and label not in seen:
            labels.append(label)
            seen.add(label)
    return labels


def count_true_label_distribution(data: List[Dict[str, Any]]) -> Tuple[Counter, int]:
    counter = Counter()
    missing_count = 0

    for item in data:
        labels = get_label_list(item.get(TRUE_LABEL_FIELD))
        if not labels:
            missing_count += 1
            continue
        counter.update(labels)

    return counter, missing_count


def count_pred_label_distribution(data: List[Dict[str, Any]]) -> Tuple[Counter, int]:
    counter = Counter()
    missing_count = 0

    for item in data:
        pred_label = normalize_label(item.get(PRED_LABEL_FIELD))
        if pred_label is None:
            missing_count += 1
            continue
        counter[pred_label] += 1

    return counter, missing_count


def process_pred(data_path: str, result_path: str) -> List[Dict[str, Any]]:
    data = read_json_or_jsonl(data_path)
    json_list = []

    for doc in data:
        content = pre_process_text(doc["content"])
        input_ids, attention_mask = token2id(content)
        pred, prob = predict(input_ids, attention_mask)
        doc[PRED_LABEL_FIELD] = pred
        doc[PRED_PROB_FIELD] = prob
        json_list.append(doc)

    with open(result_path, "w", encoding="utf-8") as f:
        import json

        json.dump(json_list, f, indent=4, ensure_ascii=False)

    return json_list


def extract_true_and_pred(
    valid_data: List[Dict[str, Any]],
) -> Tuple[List[Set[str]], List[str], int]:
    y_true_sets = []
    y_pred = []
    skip_count = 0

    for idx, item in enumerate(valid_data):
        true_labels = set(get_label_list(item.get(TRUE_LABEL_FIELD)))
        pred_label = normalize_label(item.get(PRED_LABEL_FIELD))

        if not true_labels:
            print(
                f"[空真实标签] 测试集第 {idx} 条: "
                f"labels={item.get(TRUE_LABEL_FIELD)}, model_pred={item.get(PRED_LABEL_FIELD)}"
            )
            skip_count += 1
            continue

        y_true_sets.append(true_labels)
        y_pred.append(pred_label or "")

    print(f"测试集有效评估样本数: {len(y_true_sets)}")
    print(f"测试集跳过样本数: {skip_count}")
    return y_true_sets, y_pred, skip_count


def build_indicator_matrices(
    y_true_sets: List[Set[str]],
    y_pred: List[str],
    labels: List[str],
) -> Tuple[List[List[int]], List[List[int]]]:
    label_to_idx = {label: idx for idx, label in enumerate(labels)}
    y_true_bin = []
    y_pred_bin = []

    for true_labels, pred_label in zip(y_true_sets, y_pred):
        true_row = [0] * len(labels)
        pred_row = [0] * len(labels)

        for label in true_labels:
            if label in label_to_idx:
                true_row[label_to_idx[label]] = 1

        if pred_label in label_to_idx:
            pred_row[label_to_idx[pred_label]] = 1

        y_true_bin.append(true_row)
        y_pred_bin.append(pred_row)

    return y_true_bin, y_pred_bin


def save_overall_metrics_txt(output_path, rows: List[Dict[str, Any]]):
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("========== 整体指标 ==========\n")
        for row in rows:
            f.write(f"{row['metric']}: {row['value']:.6f}\n")
            f.write(f"  计算方式: {row['calculation']}\n")
            f.write(f"  说明: {row['description']}\n")
            f.write(f"  分子: {row['numerator']}\n")
            f.write(f"  分母: {row['denominator']}\n")


def save_overall_metrics_csv(output_path, rows: List[Dict[str, Any]]):
    pd.DataFrame(rows).to_csv(output_path, index=False, encoding="utf-8-sig")


def get_model_name(model_path: str) -> str:
    normalized_path = str(model_path).strip().rstrip("/\\")
    if not normalized_path:
        return ""

    path = Path(normalized_path)
    if path.exists():
        if path.is_file():
            return path.parent.name or path.stem
        if path.name == "best_checkpoint":
            return path.parent.name
        return path.name

    parts = [part for part in normalized_path.replace("\\", "/").split("/") if part]
    if not parts:
        return normalized_path

    name = parts[-1]
    parent_name = parts[-2] if len(parts) > 1 else ""
    if name == "best_checkpoint" and len(parts) > 1:
        return parent_name
    model_file_suffixes = {".pt", ".pth", ".bin", ".safetensors", ".ckpt", ".onnx", ".pb", ".h5"}
    if Path(name).suffix.lower() in model_file_suffixes:
        return parent_name or Path(name).stem
    return name


def append_overall_metrics_xlsx(output_path: str, model_path: str, rows: List[Dict[str, Any]]):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    base_columns = ["测试时间", "模型名称", "模型路径"]
    metric_columns = [str(row["metric"]) for row in rows]
    metric_values = {str(row["metric"]): row["value"] for row in rows}
    current_row = {
        "测试时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "模型名称": get_model_name(model_path),
        "模型路径": model_path,
        **metric_values,
    }

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    if output_file.exists():
        workbook = load_workbook(output_file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1] if cell.value]
        for column in base_columns + metric_columns:
            if column not in headers:
                headers.append(column)
                sheet.cell(row=1, column=len(headers), value=column)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "overall_metrics"
        headers = base_columns + metric_columns
        sheet.append(headers)
        sheet.freeze_panes = "A2"

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    next_row = sheet.max_row + 1
    for column_idx, header in enumerate(headers, start=1):
        sheet.cell(row=next_row, column=column_idx, value=current_row.get(header, ""))

    sheet.auto_filter.ref = sheet.dimensions
    for column_idx, header in enumerate(headers, start=1):
        width = 16
        if header == "测试时间":
            width = 20
        elif header == "模型名称":
            width = 28
        elif header == "模型路径":
            width = 48
        sheet.column_dimensions[get_column_letter(column_idx)].width = width

    workbook.save(output_file)


def safe_divide(numerator: int, denominator: int) -> float:
    if denominator == 0:
        return 0.0
    return numerator / denominator


def build_per_class_acceptable_metrics(
    y_true_sets: List[Set[str]],
    y_pred: List[str],
    labels: List[str],
    true_count: Counter,
    pred_count: Counter,
) -> pd.DataFrame:
    pred_correct_count = Counter()
    acceptable_hit_count = Counter()
    selection_coverage_count = Counter()

    for true_labels, pred_label in zip(y_true_sets, y_pred):
        is_hit = pred_label in true_labels

        if is_hit:
            pred_correct_count[pred_label] += 1

        for true_label in true_labels:
            if is_hit:
                acceptable_hit_count[true_label] += 1
            if pred_label == true_label:
                selection_coverage_count[true_label] += 1

    rows = []
    for label in labels:
        gt_num = int(true_count.get(label, 0))
        pred_num = int(pred_count.get(label, 0))
        pred_correct_num = int(pred_correct_count.get(label, 0))
        acceptable_hit_num = int(acceptable_hit_count.get(label, 0))
        selection_coverage_num = int(selection_coverage_count.get(label, 0))

        rows.append(
            {
                "类别": label,
                "GT数": gt_num,
                "Pred数": pred_num,
                "预测正确数": pred_correct_num,
                "Precision": round(safe_divide(pred_correct_num, pred_num), 6),
                "可接受命中数": acceptable_hit_num,
                "可接受命中率": round(safe_divide(acceptable_hit_num, gt_num), 6),
                "选择覆盖数": selection_coverage_num,
                "选择覆盖率": round(safe_divide(selection_coverage_num, gt_num), 6),
            }
        )

    df = pd.DataFrame(rows)
    return df.sort_values(by=["GT数", "Pred数", "类别"], ascending=[False, False, True])


def build_overall_acceptable_metrics(
    y_true_sets: List[Set[str]],
    y_pred: List[str],
    per_class_df: pd.DataFrame,
    total_sample_count: int,
    empty_pred_count: int,
) -> List[Dict[str, Any]]:
    evaluated_sample_count = len(y_true_sets)
    hit_count = sum(1 for true_labels, pred_label in zip(y_true_sets, y_pred) if pred_label in true_labels)

    positive_gt_df = per_class_df[per_class_df["GT数"] > 0].sort_values(
        by=["GT数", "Pred数", "类别"], ascending=[False, False, True]
    )
    top_label_count = (
        max(1, int(len(positive_gt_df) * TOP_GT_LABEL_RATIO + 0.999999)) if not positive_gt_df.empty else 0
    )
    top_gt_df = positive_gt_df.head(top_label_count)
    scope_desc = f"GT数降序前{TOP_GT_LABEL_RATIO:.0%}类别（{top_label_count}/{len(positive_gt_df)}）"

    acceptable_rate_sum = float(top_gt_df["可接受命中率"].sum()) if not top_gt_df.empty else 0.0
    precision_sum = float(top_gt_df["Precision"].sum()) if not top_gt_df.empty else 0.0
    selection_coverage_rate_sum = float(top_gt_df["选择覆盖率"].sum()) if not top_gt_df.empty else 0.0

    rows = [
        {
            "metric": "hit_accuracy",
            "value": round(safe_divide(hit_count, evaluated_sample_count), 6),
            "numerator": hit_count,
            "denominator": evaluated_sample_count,
            "calculation": "预测标签在该样本 gt 标签集合中的样本数 / 有真实标签的评估样本数",
            "description": "整体可接受准确率；最符合“命中任一 gt 即正确”的业务规则。",
        },
        {
            "metric": "top50_macro_acceptable_hit_rate",
            "value": round(safe_divide(acceptable_rate_sum, top_label_count), 6),
            "numerator": round(acceptable_rate_sum, 6),
            "denominator": top_label_count,
            "calculation": f"{scope_desc}的可接受命中率简单平均",
            "description": "看高频类别的平均表现，避免大量低频类别主导整体指标。",
        },
        {
            "metric": "top50_macro_precision",
            "value": round(safe_divide(precision_sum, top_label_count), 6),
            "numerator": round(precision_sum, 6),
            "denominator": top_label_count,
            "calculation": f"{scope_desc}的 Precision 简单平均",
            "description": "看高频类别中模型预测该类时整体准不准；未被预测的高频类别 Precision 记为 0。",
        },
        {
            "metric": "top50_macro_selection_coverage_rate",
            "value": round(safe_divide(selection_coverage_rate_sum, top_label_count), 6),
            "numerator": round(selection_coverage_rate_sum, 6),
            "denominator": top_label_count,
            "calculation": f"{scope_desc}的选择覆盖率简单平均",
            "description": "看模型是否真的会输出这些主要类别，而不是只命中其他可接受标签。",
        },
        {
            "metric": "empty_pred_rate",
            "value": round(safe_divide(empty_pred_count, total_sample_count), 6),
            "numerator": empty_pred_count,
            "denominator": total_sample_count,
            "calculation": "预测为空的样本数 / 全部测试样本数",
            "description": "看模型输出失败或预测字段缺失的比例。",
        },
    ]
    return rows


def get_args():
    parser = ArgumentParser(description="Process multi-label testset evaluation")
    parser.add_argument("--model-path", type=str, required=True, help="Model to run evaluation")
    parser.add_argument("--eval-data", type=str, default=default_test_data_path, help="Test dataset path")
    parser.add_argument("--result-suffix", type=str, default="", help="Prediction cache suffix")
    parser.add_argument("--label-path", type=str, default=default_label_file_path, help="Label to index mapping file path")
    parser.add_argument(
        "--output-dir",
        type=str,
        default=str(SCRIPT_DIR),
        help="测试输出根目录；默认是 test_multi.py 所在目录",
    )
    parser.add_argument("--save-result-csv", type=str_to_bool, default=False, help="Whether to save CSV files")
    return parser.parse_args()


def main():
    opts = get_args()
    eval_data_path = str(Path(opts.eval_data).expanduser().resolve())
    model_path = str(Path(opts.model_path).expanduser().resolve())
    label_path = str(Path(opts.label_path).expanduser().resolve())
    output_root = Path(opts.output_dir).expanduser().resolve()
    output_root.mkdir(parents=True, exist_ok=True)

    load_model(model_file_path=model_path, label_file_path=label_path)
    run_warmup()

    if not opts.result_suffix:
        predict_save_path = f"{Path(eval_data_path).stem}_preded_multi_{datetime.now().strftime('%Y%m%d%H%M%S')}.json"
    else:
        predict_save_path = f"{Path(eval_data_path).stem}_preded_multi_{opts.result_suffix}.json"
    output_dir = output_root / Path(predict_save_path).stem
    predict_save_path = output_dir / predict_save_path

    suffix_prefix = f"{opts.result_suffix}_" if opts.result_suffix else ""

    per_class_xlsx = f"{suffix_prefix}per_class_metrics.xlsx"
    per_class_csv = f"{suffix_prefix}per_class_metrics.csv"
    overall_txt = f"{suffix_prefix}overall_metrics.txt"
    overall_csv = f"{suffix_prefix}overall_metrics.csv"
    overall_compare_xlsx_path = output_root / OVERALL_COMPARE_XLSX

    output_dir.mkdir(parents=True, exist_ok=True)

    print("========== 获取测试集预测结果 ==========")
    if predict_save_path.exists():
        valid_data = read_json_or_jsonl(predict_save_path)
        print(f"已从缓存读取测试样本数: {len(valid_data)}")
    else:
        valid_data = process_pred(eval_data_path, predict_save_path)
        print(f"从测试集执行预测并缓存: {len(valid_data)}")

    print("\n========== 统计测试集真实类别数量 ==========")
    true_count, true_missing = count_true_label_distribution(valid_data)
    print(f"测试集缺少真实标签数量: {true_missing}")

    print("\n========== 统计测试集预测类别数量 ==========")
    pred_count, pred_missing = count_pred_label_distribution(valid_data)
    print(f"测试集缺少预测标签数量: {pred_missing}")

    print("\n========== 提取真实标签集合和预测标签 ==========")
    y_true_sets, y_pred, _ = extract_true_and_pred(valid_data)

    if not y_true_sets:
        print("[错误] 没有有效测试样本，无法计算指标。")
        return

    eval_labels = sorted(set(true_count.keys()) | set(pred_count.keys()))
    print(f"\n测试集中参与评估的类别数: {len(eval_labels)}")

    print("\n========== 计算每个类别指标 ==========")
    df = build_per_class_acceptable_metrics(
        y_true_sets=y_true_sets,
        y_pred=y_pred,
        labels=eval_labels,
        true_count=true_count,
        pred_count=pred_count,
    )

    print("\n========== 计算整体指标 ==========")
    overall_rows = build_overall_acceptable_metrics(
        y_true_sets=y_true_sets,
        y_pred=y_pred,
        per_class_df=df,
        total_sample_count=len(valid_data),
        empty_pred_count=pred_missing,
    )
    for row in overall_rows:
        print(f"{row['metric']}: {row['value']:.6f}")
        print(f"  计算方式: {row['calculation']}")

    print("\n========== 每个类别统计结果 ==========")
    print(df.to_string(index=False))

    xlsx_path = output_dir / per_class_xlsx
    df.to_excel(xlsx_path, index=False, engine="openpyxl")

    if opts.save_result_csv:
        csv_path = output_dir / per_class_csv
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    overall_txt_path = output_dir / overall_txt
    save_overall_metrics_txt(
        output_path=overall_txt_path,
        rows=overall_rows,
    )

    if opts.save_result_csv:
        overall_csv_path = output_dir / overall_csv
        save_overall_metrics_csv(
            output_path=overall_csv_path,
            rows=overall_rows,
        )

    append_overall_metrics_xlsx(
        output_path=overall_compare_xlsx_path,
        model_path=model_path,
        rows=overall_rows,
    )

    print("\n========== 保存完成 ==========")
    print(f"预测缓存 JSON       : {predict_save_path}")
    print(f"每个类别指标 Excel  : {xlsx_path}")
    if opts.save_result_csv:
        print(f"每个类别指标 CSV    : {csv_path}")
    print(f"整体指标 TXT        : {overall_txt_path}")
    if opts.save_result_csv:
        print(f"整体指标 CSV        : {overall_csv_path}")
    print(f"整体指标对比 Excel  : {overall_compare_xlsx_path}")


if __name__ == "__main__":
    main()
